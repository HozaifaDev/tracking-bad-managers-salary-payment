"""Build styled Excel salary report (openpyxl)."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

EGP_FMT = '#,##0.00 "EGP"'
DATE_FMT = "yyyy-mm-dd"
HEADER_FONT = Font(bold=True)
FILL_ALT = PatternFill("solid", fgColor="FFF2F2F2")
FILL_FLAG = PatternFill("solid", fgColor="FFFFE699")
FILL_BAL_POS = PatternFill("solid", fgColor="FFFFCCCC")  # owed — red tint
FILL_BAL_ZERO = PatternFill("solid", fgColor="FFCCFFCC")
FILL_BAL_NEG = PatternFill("solid", fgColor="FFCCCCFF")  # overpaid — blue tint


def build_report(
    output_path: Path,
    summary: Dict[str, Any],
    monthly: List[Dict[str, Any]],
    session_rows: List[Dict[str, Any]],
    payments: List[Any],
    date_from: date,
    date_to: date,
) -> None:
    """Write all four sheets to output_path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    _sheet_summary(ws_sum, summary, date_from, date_to)

    ws_m = wb.create_sheet("Monthly Breakdown")
    _sheet_monthly(ws_m, monthly)

    ws_s = wb.create_sheet("Session Log")
    _sheet_sessions(ws_s, session_rows)

    ws_p = wb.create_sheet("Payment Log")
    _sheet_payments(ws_p, payments)

    wb.save(output_path)


def _sheet_summary(
    ws: Any,
    summary: Dict[str, Any],
    date_from: date,
    date_to: date,
) -> None:
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    rows = [
        ("Total Earned (All Time)", summary["total_earned_all_time"]),
        ("Total Received (All Time)", summary["total_paid_all_time"]),
        ("Current Balance Owed", summary["current_balance_owed"]),
        ("Report Generated On", generated),
        ("Date Range Covered", f"{date_from.isoformat()} → {date_to.isoformat()}"),
    ]
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    for c in ("A1", "B1"):
        ws[c].font = HEADER_FONT

    bal = float(summary["current_balance_owed"])
    for i, (label, val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2)
        if label == "Current Balance Owed":
            cell_b.value = bal
            cell_b.number_format = EGP_FMT
            if bal > 0:
                cell_b.fill = FILL_BAL_POS
            elif bal < 0:
                cell_b.fill = FILL_BAL_NEG
            else:
                cell_b.fill = FILL_BAL_ZERO
        elif "Date Range" in label or "Generated" in label:
            cell_b.value = val
        else:
            cell_b.value = float(val)
            cell_b.number_format = EGP_FMT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40


def _sheet_monthly(ws: Any, monthly: List[Dict[str, Any]]) -> None:
    headers = [
        "Salary Month",
        "Sessions Count",
        "Total Hours",
        "Expected Earnings (EGP)",
        "Cumulative Earned",
        "Cumulative Paid",
        "Running Balance",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT

    for r_idx, row in enumerate(monthly, start=2):
        row_fill = FILL_ALT if r_idx % 2 == 0 else PatternFill()
        rb = float(row["running_balance"])
        ws.cell(row=r_idx, column=1, value=row["salary_month"])
        ws.cell(row=r_idx, column=2, value=row["sessions_count"])
        ws.cell(row=r_idx, column=3, value=row["total_hours"])
        for col, key in enumerate(
            ["expected_earnings", "cumulative_earned", "cumulative_paid", "running_balance"],
            start=4,
        ):
            cell = ws.cell(row=r_idx, column=col, value=float(row[key]))
            cell.number_format = EGP_FMT
            cell.fill = FILL_BAL_POS if col == 7 and rb > 0 else row_fill
        for c in range(1, 4):
            ws.cell(row=r_idx, column=c).fill = row_fill
    _autosize_columns(ws, 1, len(headers))


def _sheet_sessions(ws: Any, session_rows: List[Dict[str, Any]]) -> None:
    headers = [
        "Date",
        "Day of Week",
        "Salary Month",
        "Event Title (raw)",
        "Category",
        "Duration (hrs)",
        "Rate (EGP/hr)",
        "Earnings (EGP)",
        "Note",
        "Flagged?",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = HEADER_FONT

    sorted_rows = sorted(session_rows, key=lambda r: (r["date"], r["title"]))
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for r_idx, row in enumerate(sorted_rows, start=2):
        d: date = row["date"]
        ws.cell(row=r_idx, column=1, value=d)
        ws.cell(row=r_idx, column=1).number_format = DATE_FMT
        ws.cell(row=r_idx, column=2, value=days[d.weekday()])
        ws.cell(row=r_idx, column=3, value=row["salary_month"])
        ws.cell(row=r_idx, column=4, value=row["title"])
        ws.cell(row=r_idx, column=5, value=row["category"])
        ws.cell(row=r_idx, column=6, value=float(row["duration_hours"]))
        ws.cell(row=r_idx, column=7, value=float(row["rate_applied"]))
        ws.cell(row=r_idx, column=7).number_format = EGP_FMT
        c_earn = ws.cell(row=r_idx, column=8, value=float(row["earnings"]))
        c_earn.number_format = EGP_FMT
        ws.cell(row=r_idx, column=9, value=row.get("note", ""))
        flagged = bool(row.get("flagged"))
        ws.cell(row=r_idx, column=10, value="Yes" if flagged else "No")
        if flagged:
            for c in range(1, 11):
                ws.cell(row=r_idx, column=c).fill = FILL_FLAG

    ws.freeze_panes = "A2"
    _autosize_columns(ws, 1, len(headers))


def _sheet_payments(ws: Any, payments: List[Any]) -> None:
    headers = ["Date Received", "Amount (EGP)", "Note", "Running Total Paid"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = HEADER_FONT

    sorted_p = sorted(payments, key=lambda p: p.date_received)
    running = 0.0
    for r_idx, p in enumerate(sorted_p, start=2):
        running = round(running + p.amount_egp, 2)
        ws.cell(row=r_idx, column=1, value=p.date_received)
        ws.cell(row=r_idx, column=1).number_format = DATE_FMT
        cell_amt = ws.cell(row=r_idx, column=2, value=p.amount_egp)
        cell_amt.number_format = EGP_FMT
        ws.cell(row=r_idx, column=3, value=p.note)
        cell_rt = ws.cell(row=r_idx, column=4, value=running)
        cell_rt.number_format = EGP_FMT
    _autosize_columns(ws, 1, len(headers))


def _autosize_columns(ws: Any, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, min(50, len(str(cell.value))))
        ws.column_dimensions[letter].width = max_len + 2
